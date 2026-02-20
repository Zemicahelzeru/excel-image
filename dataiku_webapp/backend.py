import io
import os
import posixpath
import re
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
from flask import Flask, after_this_request, jsonify, request, send_file
from werkzeug.exceptions import HTTPException

try:
    from PIL import Image
except Exception:  # pragma: no cover
    Image = None

# In Dataiku webapps, "app" is usually already provided.
# This fallback keeps the file runnable outside Dataiku for local testing.
try:
    app  # type: ignore[name-defined]
except NameError:
    app = Flask(__name__)

ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}
MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024
FIXED_IMAGE_COL = 1  # Column A
FIXED_VENDOR_COL = 4  # Column D
FIXED_MATERIAL_COL = 6  # Column F (fallback: ORIGINAL MATERIAL #)
EMU_PER_POINT = 12700
DEFAULT_ROW_HEIGHT_POINTS = 15.0


def _json_error(message, status_code=400):
    return jsonify(status="error", message=message), status_code


def _detect_ext(data):
    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return "png"
    if data.startswith(b"\xff\xd8\xff"):
        return "jpg"
    if data.startswith(b"GIF8"):
        return "gif"
    return "png"


def _normalize_ext(ext, data):
    ext = (ext or "").lower().replace(".", "")
    if ext == "jpeg":
        return "jpg"
    if ext in {"png", "jpg", "gif", "bmp", "tif", "tiff", "webp"}:
        return ext
    return _detect_ext(data)


def _safe_name(value):
    value = str(value).strip() if value not in (None, "") else "Image"
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._-") or "Image"


def _anchor_row_col(img):
    anchor = getattr(img, "anchor", None)
    if not anchor:
        return None, None

    # Some workbooks expose anchor as a plain cell reference string, e.g. "A36".
    if isinstance(anchor, str):
        return _cell_ref_to_row_col(anchor)

    from_cell = getattr(anchor, "_from", None)
    if from_cell is not None:
        row = getattr(from_cell, "row", None)
        col = getattr(from_cell, "col", None)
        if isinstance(row, int) and isinstance(col, int):
            return row + 1, col + 1

    # Fallback for marker-like anchors that directly expose row/col.
    row = getattr(anchor, "row", None)
    col = getattr(anchor, "col", None)
    if isinstance(row, int) and isinstance(col, int):
        return row + 1, col + 1

    # Final fallback: parse DrawingML-style <from><row>/<col> markers.
    row, col = _anchor_row_col_from_node(anchor, start_row=None)
    if row is not None:
        return row, col
    return None, None


def _read_up(ws, row, col):
    if not row:
        return None
    for r in range(row, 0, -1):
        value = ws.cell(r, col).value
        if value not in (None, ""):
            return value
    return None


def _next_unique_filename(base_name, ext, seen):
    candidate = "{0}.{1}".format(base_name, ext)
    if candidate not in seen:
        seen.add(candidate)
        return candidate

    counter = 2
    while True:
        candidate = "{0}_{1}.{2}".format(base_name, counter, ext)
        if candidate not in seen:
            seen.add(candidate)
            return candidate
        counter += 1


def _natural_sort_key(text):
    return [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", text)]


def _get_uploaded_file():
    file_obj = request.files.get("file")
    if not file_obj or not file_obj.filename:
        return None, None, "No file uploaded"

    ext = Path(file_obj.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        return None, None, "Please upload a valid Excel file (.xlsx or .xlsm)"

    file_bytes = file_obj.read()
    if not file_bytes:
        return None, None, "Uploaded file is empty"
    if len(file_bytes) > MAX_FILE_SIZE_BYTES:
        return None, None, "File too large. Please upload a file up to 50MB."

    return file_bytes, file_obj.filename, None


def _send_zip_response(stream, filename):
    send_kwargs = dict(
        mimetype="application/zip",
        as_attachment=True,
        conditional=False,
        etag=False,
        max_age=0,
    )
    try:
        return send_file(stream, download_name=filename, **send_kwargs)
    except TypeError:
        # Compatibility with older Flask versions.
        return send_file(stream, attachment_filename=filename, **send_kwargs)


def _normalize_label(value):
    if not isinstance(value, str):
        return ""
    return " ".join(value.strip().lower().split())


def _is_image_header_label(label):
    if not label:
        return False
    return (
        label == "image"
        or "image" in label
        or "picture" in label
        or "photo" in label
    )


def _is_vendor_header_label(label):
    return bool(label and ("vendor" in label and "material" in label))


def _is_material_header_label(label):
    if not label:
        return False
    if "vendor" in label:
        return False
    return (
        "original material" in label
        or label.startswith("material")
        or "material #" in label
    )


def _safe_folder_name(filename):
    name = (filename or "").strip()
    if not name:
        return "Excel_Images"
    # Keep original casing/spaces as much as possible while preventing invalid paths.
    name = name.replace("/", "_").replace("\\", "_")
    name = re.sub(r"[\x00-\x1f]+", "", name).strip()
    return name or "Excel_Images"


def _detect_vendor_column(ws):
    max_row = min(ws.max_row or 1, 60)
    max_col = min(ws.max_column or 1, 30)

    header_candidate = None
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            label = _normalize_label(ws.cell(row_idx, col_idx).value)
            if _is_vendor_header_label(label):
                return col_idx
            if "vendor" in label or ("material" in label and "original" not in label):
                header_candidate = header_candidate or col_idx

    if header_candidate:
        return header_candidate

    fallback_candidates = [4, 2, 3, 1]
    max_data_row = min(ws.max_row or 1, 10000)
    best_col = 4
    best_score = -1
    for col_idx in fallback_candidates:
        if col_idx > (ws.max_column or 1):
            continue
        score = 0
        for row_idx in range(2, max_data_row + 1):
            value = ws.cell(row_idx, col_idx).value
            if value not in (None, ""):
                score += 1
        if score > best_score:
            best_col = col_idx
            best_score = score
    return best_col


def _detect_material_column(ws, vendor_col):
    max_row = min(ws.max_row or 1, 60)
    max_col = min(ws.max_column or 1, 40)
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            label = _normalize_label(ws.cell(row_idx, col_idx).value)
            if _is_material_header_label(label):
                return col_idx
    # Common layout: Vendor in D, Original Material in F.
    if vendor_col and vendor_col + 2 <= (ws.max_column or 1):
        return vendor_col + 2
    return None


def _detect_layout(ws):
    max_row = min(ws.max_row or 1, 120)
    image_header_row = None
    vendor_header_row = None
    material_header_row = None

    # Strict fixed-column mode requested by user:
    # Image must be in A, vendor must be in D, optional material fallback in F.
    for row_idx in range(1, max_row + 1):
        image_label = _normalize_label(ws.cell(row_idx, FIXED_IMAGE_COL).value)
        vendor_label = _normalize_label(ws.cell(row_idx, FIXED_VENDOR_COL).value)
        material_label = _normalize_label(ws.cell(row_idx, FIXED_MATERIAL_COL).value)

        if image_header_row is None and _is_image_header_label(image_label):
            image_header_row = row_idx
        if vendor_header_row is None and _is_vendor_header_label(vendor_label):
            vendor_header_row = row_idx
        if material_header_row is None and _is_material_header_label(material_label):
            material_header_row = row_idx

    # Do not use max(header_rows): image cells can contain "Picture" text lower in the sheet
    # and would incorrectly shift every target row downward.
    if vendor_header_row is not None and vendor_header_row <= 5:
        start_row = vendor_header_row + 1
    else:
        # Avoid accidental offsets from late header-like text in data rows.
        start_row = 1

    material_col = FIXED_MATERIAL_COL if (ws.max_column or 0) >= FIXED_MATERIAL_COL else None
    return {
        "image_col": FIXED_IMAGE_COL,
        "vendor_col": FIXED_VENDOR_COL,
        "material_col": material_col,
        "start_row": start_row,
        "image_header_row": image_header_row,
        "vendor_header_row": vendor_header_row,
        "material_header_row": material_header_row,
    }


def _row_code(ws, row_idx, vendor_col, material_col):
    vendor_value = ws.cell(row_idx, vendor_col).value if vendor_col else None
    vendor_text = str(vendor_value).strip() if vendor_value not in (None, "") else ""
    vendor_label = _normalize_label(vendor_text)
    if vendor_text and not _is_vendor_header_label(vendor_label):
        return _safe_name(vendor_text), "vendor"

    material_value = ws.cell(row_idx, material_col).value if material_col else None
    material_text = str(material_value).strip() if material_value not in (None, "") else ""
    material_label = _normalize_label(material_text)
    if material_text and not _is_material_header_label(material_label):
        return _safe_name("MAT_{0}".format(material_text)), "material"

    return None, None


def _candidate_rows_for_media(ws, start_row, image_col, vendor_col, material_col):
    rows = []
    max_row = ws.max_row or start_row
    for row_idx in range(start_row, max_row + 1):
        code, _ = _row_code(ws, row_idx, vendor_col, material_col)
        image_cell = ws.cell(row_idx, image_col).value if image_col else None
        image_label = _normalize_label(image_cell)
        has_image_hint = image_label in {"image", "picture", "photo"}
        if code or has_image_hint:
            rows.append(row_idx)
    if rows:
        return rows

    # Last fallback: use rows that have any vendor/material value.
    for row_idx in range(start_row, max_row + 1):
        vendor_value = ws.cell(row_idx, vendor_col).value if vendor_col else None
        material_value = ws.cell(row_idx, material_col).value if material_col else None
        if vendor_value not in (None, "") or material_value not in (None, ""):
            rows.append(row_idx)
    return rows


def _resolve_zip_path(base_path, target):
    if not target:
        return None
    clean_target = target.replace("\\", "/")
    if clean_target.startswith("/"):
        return clean_target.lstrip("/")
    base_dir = posixpath.dirname(base_path)
    return posixpath.normpath(posixpath.join(base_dir, clean_target))


def _read_relationships(archive, rels_path):
    rels = {}
    if rels_path not in archive.namelist():
        return rels
    rel_root = ET.fromstring(archive.read(rels_path))
    ns_rel = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
    for rel in rel_root.findall("rel:Relationship", ns_rel):
        rel_id = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if rel_id and target:
            rels[rel_id] = target
    return rels


def _sheet_path_for_name(archive, sheet_name):
    workbook_path = "xl/workbook.xml"
    workbook_rels_path = "xl/_rels/workbook.xml.rels"
    if workbook_path not in archive.namelist():
        return None

    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    wb_root = ET.fromstring(archive.read(workbook_path))
    workbook_rels = _read_relationships(archive, workbook_rels_path)
    for sheet in wb_root.findall(".//main:sheets/main:sheet", ns):
        if sheet.attrib.get("name") != sheet_name:
            continue
        rid = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        target = workbook_rels.get(rid)
        return _resolve_zip_path(workbook_path, target)
    return None


def _extract_drawing_images_for_sheet(file_bytes, sheet_name, ws=None):
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    }

    entries = []
    with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as archive:
        sheet_path = _sheet_path_for_name(archive, sheet_name)
        if not sheet_path or sheet_path not in archive.namelist():
            return entries

        sheet_root = ET.fromstring(archive.read(sheet_path))
        sheet_rels_path = "{0}/_rels/{1}.rels".format(
            posixpath.dirname(sheet_path),
            posixpath.basename(sheet_path),
        )
        sheet_rels = _read_relationships(archive, sheet_rels_path)
        drawing_rids = []
        for drawing in sheet_root.findall(".//main:drawing", ns):
            rid = drawing.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if rid:
                drawing_rids.append(rid)

        item_idx = 0
        for rid in drawing_rids:
            drawing_target = sheet_rels.get(rid)
            drawing_path = _resolve_zip_path(sheet_path, drawing_target)
            if not drawing_path or drawing_path not in archive.namelist():
                continue

            drawing_root = ET.fromstring(archive.read(drawing_path))
            drawing_rels_path = "{0}/_rels/{1}.rels".format(
                posixpath.dirname(drawing_path),
                posixpath.basename(drawing_path),
            )
            drawing_rels = _read_relationships(archive, drawing_rels_path)

            drawing_anchors = (
                drawing_root.findall("xdr:twoCellAnchor", ns)
                + drawing_root.findall("xdr:oneCellAnchor", ns)
                + drawing_root.findall("xdr:absoluteAnchor", ns)
            )
            for anchor in drawing_anchors:
                local_name = _xml_local_name(getattr(anchor, "tag", "")).lower()
                row = None
                col = None
                if local_name in {"onecellanchor", "twocellanchor"}:
                    row, col = _anchor_row_col_from_node(anchor, start_row=None)
                elif local_name == "absoluteanchor":
                    pos_node = anchor.find("xdr:pos", ns)
                    if pos_node is None:
                        for child in anchor:
                            if _xml_local_name(getattr(child, "tag", "")).lower() == "pos":
                                pos_node = child
                                break
                    y_text = pos_node.attrib.get("y") if pos_node is not None else None
                    row = _row_from_y_emu(ws, y_text)

                blip = anchor.find(".//a:blip", ns)
                if blip is None:
                    continue
                embed = blip.attrib.get(
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
                )
                target = drawing_rels.get(embed)
                media_path = _resolve_zip_path(drawing_path, target)
                if not media_path or media_path not in archive.namelist():
                    continue

                data = archive.read(media_path)
                if not data:
                    continue
                item_idx += 1
                entries.append(
                    {
                        "row": row,
                        "col": col,
                        "ext": _normalize_ext(Path(media_path).suffix, data),
                        "data": data,
                        "source": "drawing:{0}".format(item_idx),
                    }
                )

    entries.sort(key=lambda item: (item.get("row") or 10**9, item.get("col") or 10**9, item["source"]))
    return entries


def _extract_sheet_related_anchor_images(file_bytes, sheet_name, start_row, ws=None):
    entries = []
    with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as archive:
        sheet_path = _sheet_path_for_name(archive, sheet_name)
        if not sheet_path or sheet_path not in archive.namelist():
            return entries

        sheet_rels_path = "{0}/_rels/{1}.rels".format(
            posixpath.dirname(sheet_path),
            posixpath.basename(sheet_path),
        )
        sheet_rels = _read_relationships(archive, sheet_rels_path)
        if not sheet_rels:
            return entries

        candidate_parts = []
        for rid, target in sheet_rels.items():
            resolved = _resolve_zip_path(sheet_path, target)
            if not resolved or resolved not in archive.namelist():
                continue
            lower = resolved.lower()
            if not lower.endswith(".xml"):
                continue
            # Parse any XML part related to the sheet because some workbooks
            # store row-anchored images in non-standard part names.
            candidate_parts.append((rid, resolved))

        seen = set()
        idx = 0
        for _rid, part_path in candidate_parts:
            if part_path in seen:
                continue
            seen.add(part_path)

            try:
                root = ET.fromstring(archive.read(part_path))
            except Exception:
                continue

            rels_path = "{0}/_rels/{1}.rels".format(
                posixpath.dirname(part_path),
                posixpath.basename(part_path),
            )
            part_rels = _read_relationships(archive, rels_path)
            if not part_rels:
                continue

            for anchor in root.iter():
                local = _xml_local_name(anchor.tag).lower()
                if local not in {"onecellanchor", "twocellanchor", "absoluteanchor"}:
                    continue

                if local in {"onecellanchor", "twocellanchor"}:
                    row, col = _anchor_row_col_from_node(anchor, start_row=start_row)
                else:
                    y_emu = _position_y_emu_from_anchor(anchor)
                    row = _row_from_y_emu(ws, y_emu)
                    col = None
                if row is None:
                    continue

                embed_rel = None
                for child in anchor.iter():
                    for attr_name, attr_value in child.attrib.items():
                        if _xml_local_name(attr_name).lower() == "embed" and attr_value:
                            embed_rel = attr_value
                            break
                    if embed_rel:
                        break
                if not embed_rel:
                    continue

                target = part_rels.get(embed_rel)
                media_path = _resolve_zip_path(part_path, target)
                if not media_path or media_path not in archive.namelist():
                    continue
                data = archive.read(media_path)
                if not data:
                    continue

                idx += 1
                entries.append(
                    {
                        "row": row,
                        "col": col,
                        "ext": _normalize_ext(Path(media_path).suffix, data),
                        "data": data,
                        "source": "sheet_related_anchor:{0}".format(idx),
                    }
                )

    entries.sort(key=lambda item: (item.get("row") or 10**9, item.get("col") or 10**9, item["source"]))
    return entries


def _extract_openpyxl_images(images):
    entries = []
    for idx, img in enumerate(images, start=1):
        row, col = _anchor_row_col(img)
        try:
            image_data = img._data()
        except Exception:
            continue
        entries.append(
            {
                "row": row,
                "col": col,
                "ext": _normalize_ext(None, image_data),
                "data": image_data,
                "source": "openpyxl:{0}".format(idx),
            }
        )
    entries.sort(key=lambda item: (item.get("row") or 10**9, item.get("col") or 10**9, item["source"]))
    return entries


def _cell_ref_to_row_col(cell_ref):
    match = re.match(r"^([A-Za-z]+)(\d+)$", cell_ref or "")
    if not match:
        return None, None
    col_letters = match.group(1).upper()
    row_idx = int(match.group(2))
    col_idx = 0
    for ch in col_letters:
        col_idx = col_idx * 26 + (ord(ch) - ord("A") + 1)
    return row_idx, col_idx


def _row_height_points(ws, row_idx):
    if ws is None:
        return DEFAULT_ROW_HEIGHT_POINTS
    default_height = (
        getattr(getattr(ws, "sheet_format", None), "defaultRowHeight", None)
        or DEFAULT_ROW_HEIGHT_POINTS
    )
    try:
        dim = ws.row_dimensions.get(row_idx)
    except Exception:
        dim = None
    height = getattr(dim, "height", None) if dim is not None else None
    if height in (None, 0):
        return float(default_height)
    try:
        return float(height)
    except Exception:
        return float(default_height)


def _row_from_y_emu(ws, y_emu):
    if ws is None or y_emu is None:
        return None
    try:
        y_value = int(y_emu)
    except Exception:
        return None
    if y_value < 0:
        y_value = 0

    max_row = (ws.max_row or 1) + 500
    cumulative = 0.0
    for row_idx in range(1, max_row + 1):
        row_height_emu = _row_height_points(ws, row_idx) * EMU_PER_POINT
        cumulative += row_height_emu
        if y_value < cumulative:
            return row_idx
    return None


def _position_y_emu_from_anchor(anchor):
    pos_node = None
    for child in anchor:
        if _xml_local_name(getattr(child, "tag", "")).lower() == "pos":
            pos_node = child
            break
    if pos_node is None:
        return None

    # xdr:pos usually has unqualified x/y attributes.
    y_value = pos_node.attrib.get("y")
    if y_value is not None:
        return y_value
    for attr_name, attr_value in pos_node.attrib.items():
        if _xml_local_name(attr_name).lower() == "y":
            return attr_value
    return None


def _anchor_row_col_from_node(anchor, start_row=None):
    row = None
    col = None

    # Find the "from" marker regardless of namespace prefix.
    from_node = None
    try:
        iterator = anchor.iter()
    except Exception:
        iterator = []
    for child in iterator:
        child_tag = getattr(child, "tag", "")
        if "from" in str(child_tag).lower():
            from_node = child
            break
    if from_node is None:
        return None, None

    for part in from_node:
        tag = getattr(part, "tag", None)
        if tag is None:
            continue
        tag_name = _xml_local_name(tag).lower()
        text_value = (part.text or "").strip()
        if tag_name == "row":
            try:
                row = int(text_value) + 1
            except Exception:
                row = None
        elif tag_name == "col":
            try:
                col = int(text_value) + 1
            except Exception:
                col = None

    # Keep exact extracted coordinates; strict mapping happens in assignment stage.
    return row, col


def _extract_dispimg_key(formula):
    if not formula:
        return None
    key_match = re.search(r'DISPIMG\(\s*"([^"]+)"', formula, flags=re.IGNORECASE)
    if not key_match:
        key_match = re.search(r"DISPIMG\(\s*'([^']+)'", formula, flags=re.IGNORECASE)
    if not key_match:
        # Some variants embed ID_* directly in longer formulas.
        generic_match = re.search(r"(ID_[A-Za-z0-9_-]+)", formula, flags=re.IGNORECASE)
        if not generic_match:
            return None
        key = (generic_match.group(1) or "").strip()
        return key or None
    key = (key_match.group(1) or "").strip()
    return key or None


def _normalize_mapping_key(key):
    return (key or "").strip().upper()


def _possible_mapping_keys(raw_value):
    text = _normalize_mapping_key(raw_value)
    if not text:
        return set()
    keys = {text}
    # Excel can wrap DISPIMG keys with extra characters. Keep stable ID_* tokens too.
    keys.update(re.findall(r"ID_[A-Z0-9_-]+", text))
    return {item for item in keys if item}


def _xml_local_name(tag):
    if "}" in tag:
        return tag.rsplit("}", 1)[-1]
    return tag


def _extract_dispimg_row_map(archive, sheet_path, image_col, start_row):
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    row_map = {}
    sheet_root = ET.fromstring(archive.read(sheet_path))
    shared_formula_by_si = {}

    for cell in sheet_root.findall(".//main:c", ns):
        cell_ref = cell.attrib.get("r")
        row_idx, col_idx = _cell_ref_to_row_col(cell_ref)
        if row_idx is None or col_idx is None:
            continue
        if col_idx != image_col:
            continue
        f_node = cell.find("main:f", ns)
        formula = ""
        if f_node is not None:
            formula = (f_node.text or "").strip()
            si = f_node.attrib.get("si")
            f_type = f_node.attrib.get("t")
            if si and formula:
                shared_formula_by_si[si] = formula
            if si and not formula and (f_type == "shared" or f_type is None):
                formula = shared_formula_by_si.get(si, "")
        key = _extract_dispimg_key(formula)
        if key:
            row_map[row_idx] = key
    return row_map


def _extract_dispimg_row_map_openpyxl(file_bytes, sheet_name, image_col, start_row):
    row_map = {}
    wb = None
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            data_only=False,
            keep_links=False,
            read_only=True,
        )
        if sheet_name not in wb.sheetnames:
            return row_map
        ws = wb[sheet_name]
        max_row = ws.max_row or 0
        for row_idx in range(1, max_row + 1):
            if start_row and row_idx < start_row:
                continue
            value = ws.cell(row_idx, image_col).value
            if not isinstance(value, str):
                continue
            key = _extract_dispimg_key(value)
            if key:
                row_map[row_idx] = key
    except Exception:
        return row_map
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
    return row_map


def _find_cellimages_part_paths(archive):
    paths = []
    seen = set()

    def _add(path):
        if path and path in archive.namelist() and path not in seen:
            seen.add(path)
            paths.append(path)

    workbook_rels = _read_relationships(archive, "xl/_rels/workbook.xml.rels")
    for target in workbook_rels.values():
        if "cellimage" in target.lower():
            part_path = _resolve_zip_path("xl/workbook.xml", target)
            _add(part_path)

    for candidate in ("xl/cellimages.xml", "xl/cellImages.xml"):
        _add(candidate)
    for name in archive.namelist():
        lower = name.lower()
        if lower.startswith("xl/") and lower.endswith(".xml") and "cell" in lower and "image" in lower:
            _add(name)
    return paths


def _find_cellimages_part_path(archive):
    paths = _find_cellimages_part_paths(archive)
    return paths[0] if paths else None


def _extract_cellimages_by_key(archive, cellimages_path, expected_keys):
    rels_path = "{0}/_rels/{1}.rels".format(
        posixpath.dirname(cellimages_path),
        posixpath.basename(cellimages_path),
    )
    rels_map = _read_relationships(archive, rels_path)
    root = ET.fromstring(archive.read(cellimages_path))
    expected_norm_keys = {
        _normalize_mapping_key(key) for key in (expected_keys or []) if _normalize_mapping_key(key)
    }
    images_by_key = {}
    record_cache = {}

    # Prefer picture nodes first for tighter key-to-media extraction.
    pic_nodes = [node for node in root.iter() if _xml_local_name(node.tag).lower() == "pic"]
    embed_nodes = []
    for node in root.iter():
        has_embed = False
        for sub in node.iter():
            for attr_name, attr_value in sub.attrib.items():
                if _xml_local_name(attr_name).lower() == "embed" and attr_value:
                    has_embed = True
                    break
            if has_embed:
                break
        if has_embed:
            embed_nodes.append(node)

    candidate_nodes = pic_nodes or embed_nodes or [
        node
        for node in root.iter()
        if _xml_local_name(node.tag).lower() in {"cellimage", "image", "onecellanchor", "twocellanchor"}
    ]

    for node in candidate_nodes:
        rel_ids = []
        for sub in node.iter():
            for attr_name, attr_value in sub.attrib.items():
                if _xml_local_name(attr_name).lower() == "embed" and attr_value:
                    rel_ids.append(attr_value)

        media_path = None
        data = None
        ext = None
        for rel_id in rel_ids:
            target = rels_map.get(rel_id)
            candidate_media_path = _resolve_zip_path(cellimages_path, target)
            if candidate_media_path and candidate_media_path in archive.namelist():
                candidate_data = archive.read(candidate_media_path)
                if candidate_data:
                    media_path = candidate_media_path
                    data = candidate_data
                    ext = _normalize_ext(Path(candidate_media_path).suffix, candidate_data)
                    break
        if not media_path or data is None:
            continue

        matched_keys = set()
        for sub in node.iter():
            for attr_value in sub.attrib.values():
                possible = _possible_mapping_keys(str(attr_value))
                matched_keys.update(possible.intersection(expected_norm_keys))
            text_value = (sub.text or "").strip()
            if text_value:
                possible = _possible_mapping_keys(text_value)
                matched_keys.update(possible.intersection(expected_norm_keys))

        # Last strict attempt: key may be embedded in longer XML strings.
        if not matched_keys:
            try:
                node_xml = _normalize_mapping_key(ET.tostring(node, encoding="unicode"))
            except Exception:
                node_xml = ""
            if node_xml:
                for expected_key in expected_norm_keys:
                    if expected_key and expected_key in node_xml:
                        matched_keys.add(expected_key)

        if len(matched_keys) != 1:
            # Accuracy-first mode: skip ambiguous or key-less nodes.
            continue

        key_norm = next(iter(matched_keys))
        record = record_cache.get(media_path)
        if record is None:
            record = {
                "data": data,
                "ext": ext or "png",
                "source": "cellimages:{0}".format(media_path),
            }
            record_cache[media_path] = record
        images_by_key[key_norm] = record

    return images_by_key


def _extract_dispimg_entries(file_bytes, sheet_name, image_col, start_row):
    entries = []
    with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as archive:
        sheet_path = _sheet_path_for_name(archive, sheet_name)
        if not sheet_path or sheet_path not in archive.namelist():
            return entries
        row_key_map = _extract_dispimg_row_map(archive, sheet_path, image_col, start_row)
        if not row_key_map:
            return entries

        cellimages_paths = _find_cellimages_part_paths(archive)
        if not cellimages_paths:
            return entries
        expected_keys = {_normalize_mapping_key(value) for value in row_key_map.values() if value}
        images_by_key = {}
        for cellimages_path in cellimages_paths:
            part_images = _extract_cellimages_by_key(archive, cellimages_path, expected_keys)
            if part_images:
                # Keep first-resolved media per key for deterministic behavior.
                for key, image in part_images.items():
                    images_by_key.setdefault(key, image)
        if not images_by_key:
            return entries

        for row_idx in sorted(row_key_map.keys()):
            norm_key = _normalize_mapping_key(row_key_map[row_idx])
            image_item = images_by_key.get(norm_key)
            if not image_item:
                continue
            entries.append(
                {
                    "row": row_idx,
                    "col": image_col,
                    "ext": image_item["ext"],
                    "data": image_item["data"],
                    "source": "dispimg:{0}".format(norm_key or row_idx),
                }
            )
    return entries


def _extract_cellimages_anchor_entries(file_bytes, image_col, start_row):
    entries = []
    with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as archive:
        cellimages_paths = _find_cellimages_part_paths(archive)
        if not cellimages_paths:
            return entries

        entry_idx = 0
        for cellimages_path in cellimages_paths:
            rels_path = "{0}/_rels/{1}.rels".format(
                posixpath.dirname(cellimages_path),
                posixpath.basename(cellimages_path),
            )
            rels_map = _read_relationships(archive, rels_path)
            if not rels_map:
                continue

            root = ET.fromstring(archive.read(cellimages_path))
            for anchor in root.iter():
                anchor_local = _xml_local_name(anchor.tag).lower()
                if anchor_local not in {"onecellanchor", "twocellanchor"}:
                    continue

                row, col = _anchor_row_col_from_node(anchor, start_row=start_row)
                if row is None:
                    continue

                embed_rel = None
                for child in anchor.iter():
                    for attr_name, attr_value in child.attrib.items():
                        if _xml_local_name(attr_name).lower() == "embed" and attr_value:
                            embed_rel = attr_value
                            break
                    if embed_rel:
                        break
                if not embed_rel:
                    continue

                target = rels_map.get(embed_rel)
                media_path = _resolve_zip_path(cellimages_path, target)
                if not media_path or media_path not in archive.namelist():
                    continue
                data = archive.read(media_path)
                if not data:
                    continue

                entry_idx += 1
                entries.append(
                    {
                        "row": row,
                        "col": col,
                        "ext": _normalize_ext(Path(media_path).suffix, data),
                        "data": data,
                        "source": "cellimages_anchor:{0}".format(entry_idx),
                    }
                )

    entries.sort(key=lambda item: (item.get("row") or 10**9, item.get("col") or 10**9, item["source"]))
    return entries


def _extract_dispimg_row_keys(file_bytes, sheet_name, image_col, start_row):
    row_map = {}
    with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as archive:
        sheet_path = _sheet_path_for_name(archive, sheet_name)
        if sheet_path and sheet_path in archive.namelist():
            row_map = _extract_dispimg_row_map(archive, sheet_path, image_col, start_row)

    # Fallback parser for workbooks where XML shared-formula parsing misses DISPIMG rows.
    openpyxl_row_map = _extract_dispimg_row_map_openpyxl(
        file_bytes,
        sheet_name,
        image_col,
        start_row,
    )
    if openpyxl_row_map:
        for row_idx, key in openpyxl_row_map.items():
            row_map.setdefault(row_idx, key)
    return row_map


def _collect_target_rows(ws, start_row, vendor_col, material_col):
    rows = []
    max_row = ws.max_row or start_row
    for row_idx in range(start_row, max_row + 1):
        code, _ = _row_code(ws, row_idx, vendor_col, material_col)
        if code:
            rows.append(row_idx)
    return rows


def _assign_entries_to_rows(target_rows, entries, image_col):
    diagnostics = {
        "strategy": "strict_coordinate",
        "exact_row_matches": 0,
        "strict_col_matches": 0,
        "missing_rows": [],
    }
    if not target_rows or not entries:
        diagnostics["missing_rows"] = sorted(target_rows or [])
        return [], diagnostics

    # Strict coordinate mapping: row-id -> image entry.
    strict_map = {entry["row"]: entry for entry in entries if entry.get("row") is not None}

    mapped = []
    missing_rows = []
    for row_idx in sorted(target_rows):
        entry = strict_map.get(row_idx)
        if entry is None:
            missing_rows.append(row_idx)
        else:
            mapped.append((row_idx, entry))

    diagnostics["exact_row_matches"] = len(mapped)
    diagnostics["strict_col_matches"] = len(mapped)
    diagnostics["missing_rows"] = missing_rows
    return mapped, diagnostics


def _maybe_upscale_image(data, ext, scale_factor=3):
    if Image is None:
        return data, ext, False

    try:
        with Image.open(io.BytesIO(data)) as img:
            # Upscale only when image is visually small.
            if max(img.size) >= 300:
                return data, ext, False

            resampling = getattr(getattr(Image, "Resampling", Image), "LANCZOS", Image.BICUBIC)
            resized = img.resize((img.width * scale_factor, img.height * scale_factor), resampling)
            output = io.BytesIO()
            # Save as PNG to avoid additional quality loss from JPEG re-encoding.
            resized.save(output, format="PNG")
            return output.getvalue(), "png", True
    except Exception:
        return data, ext, False


def _extract_media_images(file_bytes):
    media = []
    with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as archive:
        media_files = [
            name for name in archive.namelist() if name.lower().startswith("xl/media/")
        ]
        media_files.sort(key=_natural_sort_key)
        for media_name in media_files:
            try:
                data = archive.read(media_name)
            except Exception:
                continue
            if not data:
                continue
            ext = _normalize_ext(Path(media_name).suffix, data)
            media.append(
                {
                    "source": media_name,
                    "ext": ext,
                    "data": data,
                }
            )
    return media


@app.errorhandler(Exception)
def _handle_unexpected_error(exc):
    if isinstance(exc, HTTPException):
        return exc
    app.logger.exception("Unhandled webapp backend exception")
    return _json_error("Backend error while processing request: {0}".format(exc), 500)


@app.route("/health")
def health():
    return jsonify(status="ok")


@app.route("/backend/health")
def health_backend_alias():
    return jsonify(status="ok")


@app.route("/get_sheets", methods=["POST"])
def get_sheets():
    file_bytes, _filename, error = _get_uploaded_file()
    if error:
        return _json_error(error, 400)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
    except Exception as exc:
        return _json_error("Could not read Excel file: {0}".format(exc), 400)

    if not sheet_names:
        return _json_error("No sheets found in the Excel file", 400)

    return jsonify(status="ok", sheets=sheet_names)


@app.route("/extract_images", methods=["POST"])
def extract_images():
    file_bytes, original_filename, error = _get_uploaded_file()
    if error:
        return _json_error(error, 400)

    sheet_name = (request.form.get("sheet_name") or "").strip()
    if not sheet_name:
        return _json_error("Missing sheet_name", 400)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, keep_links=False)
    except Exception as exc:
        return _json_error("Could not open workbook: {0}".format(exc), 400)

    if sheet_name not in wb.sheetnames:
        wb.close()
        return _json_error('Sheet "{0}" not found in workbook'.format(sheet_name), 400)

    ws = wb[sheet_name]
    layout = _detect_layout(ws)
    image_col = layout["image_col"]
    vendor_col = layout["vendor_col"]
    material_col = layout["material_col"]
    start_row = layout["start_row"]
    target_rows = _collect_target_rows(ws, start_row, vendor_col, material_col)

    openpyxl_images = list(getattr(ws, "_images", []) or [])
    openpyxl_entries = _extract_openpyxl_images(openpyxl_images)
    drawing_entries = _extract_drawing_images_for_sheet(file_bytes, sheet_name, ws=ws)
    related_anchor_entries = _extract_sheet_related_anchor_images(file_bytes, sheet_name, start_row, ws=ws)
    dispimg_entries = _extract_dispimg_entries(file_bytes, sheet_name, image_col, start_row)
    cellimages_anchor_entries = _extract_cellimages_anchor_entries(file_bytes, image_col, start_row)
    dispimg_row_keys = _extract_dispimg_row_keys(file_bytes, sheet_name, image_col, start_row)
    media_images = _extract_media_images(file_bytes)
    unique_dispimg_keys = len(
        {
            _normalize_mapping_key(value)
            for value in dispimg_row_keys.values()
            if _normalize_mapping_key(value)
        }
    )

    extracted_count = 0
    skipped_count = 0
    skipped_reasons = []
    seen_filenames = set()
    extraction_mode = "none"
    upscaled_count = 0
    mapping_info = {
        "strategy": "none",
        "exact_row_matches": 0,
        "strict_col_matches": 0,
        "missing_rows": [],
    }

    excel_name_no_ext = Path(original_filename).stem if original_filename else "Excel_Images"
    root_folder = _safe_folder_name(excel_name_no_ext)
    download_name = "{0}.zip".format(root_folder)

    temp_file = tempfile.NamedTemporaryFile(
        prefix="dataiku_excel_images_",
        suffix=".zip",
        delete=False,
    )
    zip_path = temp_file.name
    temp_file.close()

    try:
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zip_file:
            source_entries = []
            if dispimg_entries:
                source_entries = dispimg_entries
                extraction_mode = "dispimg_cellimages"
            elif related_anchor_entries:
                source_entries = related_anchor_entries
                extraction_mode = "sheet_related_anchor"
            elif cellimages_anchor_entries:
                source_entries = cellimages_anchor_entries
                extraction_mode = "cellimages_anchor"
            elif drawing_entries:
                source_entries = drawing_entries
                extraction_mode = "drawing_anchor"
            elif openpyxl_entries:
                source_entries = openpyxl_entries
                extraction_mode = "openpyxl_anchor"

            image_cache = {}

            def _prepared_image(entry):
                cache_key = entry.get("source") or id(entry)
                if cache_key in image_cache:
                    return image_cache[cache_key]
                new_data, new_ext, did_upscale = _maybe_upscale_image(entry["data"], entry["ext"], scale_factor=3)
                image_cache[cache_key] = (new_data, new_ext, did_upscale)
                return image_cache[cache_key]

            def _write_entry_for_row(row_idx, entry):
                nonlocal extracted_count, upscaled_count
                safe_code, code_source = _row_code(ws, row_idx, vendor_col, material_col)
                if not safe_code:
                    safe_code = "Row_{0}".format(row_idx)
                if code_source == "material":
                    skipped_reasons.append(
                        "Row {0}: vendor missing, used material fallback MAT_.".format(row_idx)
                    )
                out_data, out_ext, did_upscale = _prepared_image(entry)
                if did_upscale:
                    upscaled_count += 1
                filename = _next_unique_filename(safe_code, out_ext, seen_filenames)
                zip_file.writestr("{0}/{1}".format(root_folder, filename), out_data)
                extracted_count += 1

            # Strict behavior: same row only (vendor/material row N -> image row N).
            if source_entries and target_rows:
                mapped_rows, mapping_info = _assign_entries_to_rows(target_rows, source_entries, image_col)
                for row_idx, entry in mapped_rows:
                    _write_entry_for_row(row_idx, entry)

                # Report anchors that were found but do not belong to vendor/material rows.
                target_set = set(target_rows)
                non_target_anchor_rows = sorted(
                    {
                        entry.get("row")
                        for entry in source_entries
                        if entry.get("row") is not None and entry.get("row") not in target_set
                    }
                )
                if non_target_anchor_rows:
                    preview_extra = ",".join(str(r) for r in non_target_anchor_rows[:20])
                    skipped_reasons.append(
                        "Found image anchors on non-target rows: {0}{1}".format(
                            preview_extra,
                            "..." if len(non_target_anchor_rows) > 20 else "",
                        )
                    )

                missing_rows = mapping_info.get("missing_rows") or []
                if missing_rows:
                    extraction_mode = "strict_row_locked_missing_rows"
                    preview = ",".join(str(r) for r in missing_rows[:20])
                    skipped_reasons.append(
                        "Strict row-only mapping missing image rows: {0}{1}. "
                        "No order-based fallback is allowed.".format(
                            preview,
                            "..." if len(missing_rows) > 20 else "",
                        )
                    )
                    skipped_count += len(missing_rows)

            elif target_rows:
                extraction_mode = "strict_row_locked_no_row_anchors"
                mapping_info["strategy"] = "strict_only_no_source_entries"
                mapping_info["exact_row_matches"] = 0
                mapping_info["strict_col_matches"] = 0
                mapping_info["missing_rows"] = sorted(target_rows)
                skipped_count += len(target_rows)
                skipped_reasons.append(
                    "No row-anchored images found for target rows. "
                    "Strict mode forbids row-order or code-order guessing. [build: strict-v8]"
                )
                skipped_reasons.append(
                    "Diagnostics: DISPIMG rows={0}, unique DISPIMG keys={1}, "
                    "media images={2}, drawing anchors={3}, openpyxl anchors={4}.".format(
                        len(dispimg_row_keys),
                        unique_dispimg_keys,
                        len(media_images),
                        len(drawing_entries),
                        len(openpyxl_entries),
                    )
                )

            # If no vendor/material rows were found, still export discovered images.
            elif source_entries:
                for idx, entry in enumerate(source_entries, start=1):
                    safe_code = "Image_{0}".format(idx)
                    out_data, out_ext, did_upscale = _prepared_image(entry)
                    if did_upscale:
                        upscaled_count += 1
                    filename = _next_unique_filename(safe_code, out_ext, seen_filenames)
                    zip_file.writestr("{0}/{1}".format(root_folder, filename), out_data)
                    extracted_count += 1

            elif media_images and not target_rows:
                extraction_mode = "xlsx_media_no_target_rows"
                for idx, media in enumerate(media_images, start=1):
                    safe_code = "Image_{0}".format(idx)
                    out_data, out_ext, did_upscale = _maybe_upscale_image(media["data"], media["ext"], scale_factor=3)
                    if did_upscale:
                        upscaled_count += 1
                    filename = _next_unique_filename(safe_code, out_ext, seen_filenames)
                    zip_file.writestr("{0}/{1}".format(root_folder, filename), out_data)
                    extracted_count += 1

            summary_lines = [
                "Excel Image Extraction Summary",
                "==============================",
                "Generated at: {0}Z".format(datetime.utcnow().isoformat()),
                "Sheet: {0}".format(sheet_name),
                "Workbook file: {0}".format(original_filename or ""),
                "Output root folder: {0}".format(root_folder),
                "Extraction mode: {0}".format(extraction_mode),
                "Detected image column: {0}".format(image_col),
                "Detected vendor column: {0}".format(vendor_col),
                "Detected material column: {0}".format(material_col or "none"),
                "Data start row: {0}".format(start_row),
                "Vendor/material target rows: {0}".format(len(target_rows)),
                "DISPIMG/cellimages entries: {0}".format(len(dispimg_entries)),
                "Sheet-related row-anchored entries: {0}".format(len(related_anchor_entries)),
                "Cellimages row-anchored entries: {0}".format(len(cellimages_anchor_entries)),
                "DISPIMG formula rows in image column: {0}".format(len(dispimg_row_keys)),
                "Openpyxl anchored images: {0}".format(len(openpyxl_entries)),
                "Drawing anchored images: {0}".format(len(drawing_entries)),
                "XLSX media items: {0}".format(len(media_images)),
                "Upscaled images (3x): {0}".format(upscaled_count),
                "Mapping strategy: {0}".format(mapping_info.get("strategy")),
                "Mapping exact row matches: {0}".format(mapping_info.get("exact_row_matches")),
                "Mapping strict image-column matches: {0}".format(mapping_info.get("strict_col_matches")),
                "Mapping missing rows: {0}".format(len(mapping_info.get("missing_rows") or [])),
                "Extracted images: {0}".format(extracted_count),
                "Skipped images: {0}".format(skipped_count),
                "",
                "Rules:",
                "- Strict same-row mapping: vendor/material row N uses image row N.",
                "- Row mapping starts after detected header rows.",
                "- No nearest-row guessing and no offset guessing.",
                "- No row-order fallback and no code-group fallback are allowed.",
                "- Images anchored to non-target rows are never reassigned to a different row.",
                "- Rows without anchored images are reported as missing.",
                "- Preferred name is Vendor Material from detected vendor column.",
                "- If Vendor is empty and Original Material exists, file uses MAT_<material>.",
                "",
            ]
            if skipped_reasons:
                summary_lines.append("Skipped details:")
                summary_lines.extend("- {0}".format(reason) for reason in skipped_reasons)

            zip_file.writestr("{0}/summary.txt".format(root_folder), "\n".join(summary_lines))
    except Exception:
        try:
            os.remove(zip_path)
        except OSError:
            pass
        wb.close()
        raise

    wb.close()

    if extracted_count == 0:
        try:
            os.remove(zip_path)
        except OSError:
            pass
        if skipped_reasons:
            error_message = " | ".join(skipped_reasons[:3])
        else:
            error_message = "No images were extracted. Ensure images are in Column A and not empty."
        return _json_error(
            error_message,
            400,
        )

    @after_this_request
    def _cleanup_temp_file(response):
        try:
            os.remove(zip_path)
        except OSError:
            pass
        return response

    return _send_zip_response(zip_path, download_name)
