from __future__ import annotations

import io
import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple

import openpyxl
from flask import Flask, jsonify, render_template, request, send_file

app = Flask(__name__)

ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}
MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024


@app.route("/")
def index():
    return render_template("index.html")


def _json_error(message: str, status_code: int = 400):
    return jsonify(status="error", message=message), status_code


def _detect_ext(data: bytes) -> str:
    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return "png"
    if data.startswith(b"\xff\xd8\xff"):
        return "jpg"
    if data.startswith(b"GIF8"):
        return "gif"
    return "png"


def _safe_name(value) -> str:
    value = str(value).strip() if value not in (None, "") else "Image"
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._-") or "Image"


def _anchor_row_col(img) -> Tuple[Optional[int], Optional[int]]:
    anchor = getattr(img, "anchor", None)
    if anchor:
        from_cell = getattr(anchor, "_from", None)
        if from_cell:
            return from_cell.row + 1, from_cell.col + 1
    return None, None


def _read_up(ws, row: Optional[int], col: int):
    if not row:
        return None
    for r in range(row, 0, -1):
        value = ws.cell(r, col).value
        if value not in (None, ""):
            return value
    return None


def _next_unique_filename(base_name: str, ext: str, seen: set[str]) -> str:
    candidate = f"{base_name}.{ext}"
    if candidate not in seen:
        seen.add(candidate)
        return candidate

    counter = 2
    while True:
        candidate = f"{base_name}_{counter}.{ext}"
        if candidate not in seen:
            seen.add(candidate)
            return candidate
        counter += 1


def _get_uploaded_file():
    file_obj = request.files.get("file")
    if not file_obj or not file_obj.filename:
        return None, "No file uploaded"

    ext = Path(file_obj.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        return None, "Please upload a valid Excel file (.xlsx or .xlsm)"

    file_bytes = file_obj.read()
    if not file_bytes:
        return None, "Uploaded file is empty"
    if len(file_bytes) > MAX_FILE_SIZE_BYTES:
        return None, "File too large. Please upload a file up to 50MB."

    return file_bytes, None


@app.route("/health")
@app.route("/backend/health")
def health():
    return jsonify(status="ok")


@app.route("/get_sheets", methods=["POST"])
@app.route("/backend/get_sheets", methods=["POST"])
def get_sheets():
    file_bytes, error = _get_uploaded_file()
    if error:
        return _json_error(error, 400)

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
        sheet_names = wb.sheetnames
        wb.close()
    except Exception as exc:
        return _json_error(f"Could not read Excel file: {exc}", 400)

    if not sheet_names:
        return _json_error("No sheets found in the Excel file", 400)

    return jsonify(status="ok", sheets=sheet_names)


@app.route("/extract_images", methods=["POST"])
@app.route("/backend/extract_images", methods=["POST"])
def extract_images():
    file_bytes, error = _get_uploaded_file()
    if error:
        return _json_error(error, 400)

    sheet_name = (request.form.get("sheet_name") or "").strip()
    if not sheet_name:
        return _json_error("Missing sheet_name", 400)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        return _json_error(f"Could not open workbook: {exc}", 400)

    if sheet_name not in wb.sheetnames:
        wb.close()
        return _json_error(f'Sheet "{sheet_name}" not found in workbook', 400)

    ws = wb[sheet_name]
    images = list(getattr(ws, "_images", []) or [])

    out = io.BytesIO()
    extracted_count = 0
    skipped_count = 0
    skipped_reasons: list[str] = []
    seen_filenames: set[str] = set()

    with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for idx, img in enumerate(images, start=1):
            row, col = _anchor_row_col(img)
            if col != 1:
                skipped_count += 1
                skipped_reasons.append(
                    f"Image #{idx}: skipped (not in Column A). Found column={col or 'unknown'}."
                )
                continue

            vendor = _read_up(ws, row, 4)
            safe_vendor = _safe_name(vendor) if vendor else f"Row_{row or idx}"

            try:
                image_data = img._data()
            except Exception as exc:
                skipped_count += 1
                skipped_reasons.append(f"Image #{idx}: could not read image data ({exc}).")
                continue

            ext = _detect_ext(image_data)
            filename = _next_unique_filename(safe_vendor, ext, seen_filenames)
            zf.writestr(f"images/{filename}", image_data)
            extracted_count += 1

        summary_lines = [
            "Excel Image Extraction Summary",
            "==============================",
            f"Generated at: {datetime.utcnow().isoformat()}Z",
            f"Sheet: {sheet_name}",
            f"Total images found in sheet: {len(images)}",
            f"Extracted images: {extracted_count}",
            f"Skipped images: {skipped_count}",
            "",
            "Rules:",
            "- Images are extracted only when anchored in Column A.",
            "- File names are based on nearest non-empty value above/in Column D.",
            "",
        ]
        if skipped_reasons:
            summary_lines.append("Skipped details:")
            summary_lines.extend(f"- {reason}" for reason in skipped_reasons)

        zf.writestr("summary.txt", "\n".join(summary_lines))

    wb.close()
    out.seek(0)

    if extracted_count == 0:
        return _json_error(
            "No images were extracted. Ensure images are in Column A and not empty.",
            400,
        )

    return send_file(
        out,
        mimetype="application/zip",
        as_attachment=True,
        download_name="images.zip",
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
